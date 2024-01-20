# -*- coding: utf-8 -*-
"""
Created on Mon Jan 14 10:03:38 2019

@author: elisn
"""
# import difflib
import datetime

import numpy as np
import openpyxl
from numpy import atleast_1d as arr

weekdd = ["0{0}".format(i) for i in range(1, 10)] + [str(i) for i in range(10, 53)]


def score_name(name1, name2, aao=False):
    """Hur lika är name1 de olika elementen i name2?
    Returnerar array name1 * name2
    aao = True -> ersätter Å, Ä, Ö med A, A, O"""

    def swe(s):
        s = s.replace("Å", "A")
        s = s.replace("Ä", "A")
        s = s.replace("Ö", "O")
        return s

    if aao:
        name1 = [swe(n) for n in name1]
        name2 = [swe(n) for n in name2]

    import difflib

    score = np.zeros((len(name1), len(name2)))

    for n, n1 in enumerate(name1):
        score[n, :] = [difflib.SequenceMatcher(None, b, n1).ratio() for b in name2]

    return score


def splitnonalpha(s):
    pos = 1
    while pos < len(s) and s[pos].isalpha():
        pos += 1
    return (s[:pos], s[pos:])


def read_excel_table(file, worksheet=None, headers=[], srow=1):
    """Reads any excel data table into list. Each row is stored as a dictionary
    in the list, with it's values stored using the column names as keys. Assumes
    the excel file contains headers in the first row.
    Inputs:
        file - complete file path
        worksheet - name of worksheet to read
        headers - list of headers, if different from headers specified in sheet
        srow - starting row
    Outputs:
        data - list of dictionaries, one for each data row
        fields - dictionary keys, same as headers
    """

    wb = openpyxl.load_workbook(file)

    if worksheet is None:  # select first worksheet
        ws = wb.worksheets[0]
    else:  # select named worksheet
        ws = wb[worksheet]

    # first row contains header
    if headers == [] or headers.__len__() != ws.max_column:
        fields = [ws.cell(srow, i).value for i in range(1, ws.max_column + 1)]
    else:
        fields = headers

    # Header may span multiple rows, increment row counter until we find
    # first non-empty cell
    srow = srow + 1
    while ws.cell(srow, 1).value is None:
        srow = srow + 1

    # read all rows as dicts into list
    data = []
    for i in range(srow, ws.max_row + 1):
        d = {}
        for j in range(1, ws.max_column + 1):
            d[fields[j - 1]] = ws.cell(i, j).value
        data.append(d)

    wb.close()

    return (data, fields)


def new_zero_dict(params):
    """Create a new dictionary with zero/empty fields taken from params"""
    d = {}
    for p in params:
        # d[p[0]] = p[1]()
        if p[1] is list:
            d[p[0]] = None
        else:
            d[p[0]] = p[1]()
    return d


def new_duplicate_dict(d):
    """Create a duplicate of the given dictionary"""
    d1 = {}
    for p in d.keys():
        d1[p] = d[p]
    return d1


def find_str(s, l):
    """Find index of occurence of string s in list l"""
    idx = 0
    while idx < l.__len__() and l[idx] != s:
        idx = idx + 1
    return idx


def format_comment(comment, max_line_length):
    """Break up a string into several lines of given maximum length"""
    # accumulated line length
    ACC_length = 0
    words = comment.split(" ")
    formatted_comment = ""
    for word in words:
        # if ACC_length + len(word) and a space is <= max_line_length
        if ACC_length + (len(word) + 1) <= max_line_length:
            # append the word and a space
            formatted_comment = formatted_comment + word + " "
            # length = length + length of word + length of space
            ACC_length = ACC_length + len(word) + 1
        else:
            # append a line break, then the word and a space
            formatted_comment = formatted_comment + "\n" + word + " "
            # reset counter of length to the length of a word and a space
            ACC_length = len(word) + 1
    return formatted_comment


def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3


def str_to_date(strdate):
    """Take a string with a date and return datetime object
    Allowed formats:
        'YYYYMMDD'
        'YYYY-MM-DD'
        'YYYYMMDD:HH'
        'YYYY-MM-DD:HH'
    """
    year = int(strdate[0:4])
    if strdate[4] == "-":
        month = int(strdate[5:7])
        day = int(strdate[8:10])
        idx = 10
    else:
        month = int(strdate[4:6])
        day = int(strdate[6:8])
        idx = 8
    if strdate.__len__() > idx:
        hour = int(strdate[idx + 1 : idx + 3])
        return datetime.datetime(year, month, day, hour)
    else:
        return datetime.datetime(year, month, day)


def week_to_date(weekstr):
    """Given week in format 'YYYY:WW' return datetime object with date for start of week"""
    year = int(weekstr[0:4])
    week = int(weekstr[5:7])
    return datetime.datetime(year, 1, 1) + datetime.timedelta(days=(week - 1) * 7)


def weeks_in_interval(start, end):
    """Given start and end time in datetime format, return all weeks partially
    covered by the dates as 'YYYY:WW'"""
    weeks = []
    while start < end:
        if date_to_week(start) not in weeks:  # last week is 8 days, prevent it from being entered twice
            weeks.append(date_to_week(start))
        start += datetime.timedelta(days=7)
    if date_to_week(end) not in weeks:
        weeks.append(date_to_week(end))
    return weeks


def week_to_range(week, year):
    """Return first and last hour in week in given year"""

    start = datetime.datetime(year, 1, 1) + datetime.timedelta(days=7 * (week - 1))
    end = start + datetime.timedelta(days=7) + datetime.timedelta(seconds=-3600)
    if (end + datetime.timedelta(days=7)).year > year:
        end = datetime.datetime(year, 12, 31, 23)

    return (start, end)


def date_to_week(date):
    """Given datetime, find the week and return string with format 'YYYY:WW'"""
    sdate = datetime.datetime(date.year, 1, 1) - datetime.timedelta(hours=1)
    # increment one week at a time, check if sdate exceeds date
    widx = 0
    while sdate < date:
        sdate += datetime.timedelta(days=7)
        widx += 1
        # one year is 52*7+1 days, by doing this we include the last day of the
    # year in week 52
    if widx == 53:
        widx = 52
    return "{0}:{1}".format(str(date.year), weekdd[widx - 1])


def increment_week(weekstr):
    """Given week in format 'YYYY:WW', return next week"""
    year = int(weekstr.split(":")[0])
    week = int(weekstr.split(":")[1])
    if week < 52:
        return "{0}:{1}".format(year, weekdd[week])
    else:
        return "{0}:01".format(year + 1)


def decrement_week(weekstr):
    """Given week in format 'YYYY:WW', return previous week"""
    year = int(weekstr.split(":")[0])
    week = int(weekstr.split(":")[1])
    if week > 1:
        return "{0}:{1}".format(year, weekdd[week - 2])
    else:
        return "{0}:52".format(year - 1)


def c_trans(in1, in2=None, rt90=False):
    """Transform between lat/lon and SWEREF99TM or RT90"""

    if in2 is None:
        in1, in2 = in1[:, 0], in1[:, 1]
        return_matrix = True
    else:
        return_matrix = False

    in1 = arr(in1).astype(float)
    in2 = arr(in2).astype(float)
    ind = ((np.isnan(in1)) | (in1 == -9999)) | ((np.isnan(in2)) | (in2 == -9999))  # beakta ej
    out1 = np.nan * in1
    out2 = np.nan * in2
    in1 = in1[~ind]
    in2 = in2[~ind]

    from numpy import arcsin, arctan, arctanh, cos, cosh, pi, sin, sinh, tan

    asin, atan, atanh = arcsin, arctan, arctanh

    if np.nansum(in1) < np.nansum(in2):  # in1 ska vara nordlig koordinat
        in1, in2 = in2, in1

    a = 6378137.0  # semi-major axis of the ellipsoid
    f = 1 / 298.257222101  # flattening of the ellipsoid
    k0 = 0.9996  # scale factor along the central meridian
    FN = 0.0  # false northing
    FE = 500000.0  # false easting
    lambda0 = 15.0 * pi / 180  # central meridian

    if rt90:  # RT90 (GRS80)
        FE = 1500064.274
        FN = -667.711
        lambda0 = (15.0 + 48.0 / 60 + 22.624306 / 3600) * pi / 180
        k0 = 1.00000561024

    e = (f * (2 - f)) ** 0.5
    n = f / (2 - f)
    a_hat = a / (1 + n) * (1 + n**2 / 4 + n**4 / 64)

    if np.nanmax(((in1, in2))) > 1000:
        # x, y -> lat, long
        xi = (in1 - FN) / (k0 * a_hat)
        eta = (in2 - FE) / (k0 * a_hat)
        delta1 = n / 2 - 2 * n**2 / 3 + 37 * n**3 / 96 - n**4 / 360
        delta2 = n**2 / 48 + n**3 / 15 - 437 * n**4 / 1440
        delta3 = 17 * n**3 / 480 - 37 * n**4 / 840
        delta4 = 4397 * n**4 / 161280
        xi2 = (
            xi
            - delta1 * sin(2 * xi) * cosh(2 * eta)
            - delta2 * sin(4 * xi) * cosh(4 * eta)
            - delta3 * sin(6 * xi) * cosh(6 * eta)
            - delta4 * sin(8 * xi) * cosh(8 * eta)
        )
        eta2 = (
            eta
            - delta1 * cos(2 * xi) * sinh(2 * eta)
            - delta2 * cos(4 * xi) * sinh(4 * eta)
            - delta3 * cos(6 * xi) * sinh(6 * eta)
            - delta4 * cos(8 * xi) * sinh(8 * eta)
        )
        phi = asin(sin(xi2) / cosh(eta2))
        lambda1 = atan(sinh(eta2) / cos(xi2))
        A = e**2 + e**4 + e**6 + e**8
        B = -1.0 / 6 * (7 * e**4 + 17 * e**6 + 30 * e**8)
        C = 1.0 / 120 * (224 * e**6 + 889 * e**8)
        D = -1.0 / 1260 * 4279 * e**8
        out1[~ind] = (
            180.0
            / pi
            * (phi + sin(phi) * cos(phi) * (A + B * (sin(phi)) ** 2 + C * (sin(phi)) ** 4 + D * (sin(phi)) ** 6))
        )
        out2[~ind] = 180.0 / pi * (lambda1 + lambda0)
    else:
        # lat, long -> x, y
        phi = in1 * pi / 180
        lambda1 = in2 * pi / 180
        A = e**2
        B = (5 * e**4 - e**6) / 6
        C = (104 * e**6 - 45 * e**8) / 120
        D = 1237 * e**8 / 1260
        phi2 = phi - sin(phi) * cos(phi) * (A + B * (sin(phi)) ** 2 + C * (sin(phi)) ** 4 + D * (sin(phi)) ** 6)
        lambda2 = lambda1 - lambda0
        xi = atan(tan(phi2) / cos(lambda2))
        eta = atanh(cos(phi2) * sin(lambda2))
        beta1 = n / 2 - 2 * n**2 / 3 + 5 * n**3 / 16 + 41 * n**4 / 180
        beta2 = 13 * n**2 / 48 - 3 * n**3 / 5 + 557 * n**4 / 1440
        beta3 = 61 * n**3 / 240 - 103 * n**4 / 140
        beta4 = 49561 * n**4 / 161280
        out1[~ind] = (
            k0
            * a_hat
            * (
                xi
                + beta1 * sin(2 * xi) * cosh(2 * eta)
                + beta2 * sin(4 * xi) * cosh(4 * eta)
                + beta3 * sin(6 * xi) * cosh(6 * eta)
                + beta4 * sin(8 * xi) * cosh(8 * eta)
            )
            + FN
        )
        out2[~ind] = (
            k0
            * a_hat
            * (
                eta
                + beta1 * cos(2 * xi) * sinh(2 * eta)
                + beta2 * cos(4 * xi) * sinh(4 * eta)
                + beta3 * cos(6 * xi) * sinh(6 * eta)
                + beta4 * cos(8 * xi) * sinh(8 * eta)
            )
            + FE
        )

    if return_matrix:
        return np.column_stack((out1, out2))
    else:
        if len(out1) == 1:
            return out1[0], out2[0]
        else:
            return out1, out2


if __name__ == "__main__":
    #    start = datetime.datetime(2018,1,1)
    #    end = datetime.datetime(2018,12,31,23)
    #    weeks_in_interval(start,end)
    #    print('hej')

    week_to_range(1, 2018)
